# from sklearn import svm
# import numpy as np
# import matplotlib.pyplot as plt

# #from sklearn.linear_model import SGDClassifier

# # we create 40 separable points
# rng = np.random.RandomState(0)
# n_samples_1 = 1000
# n_samples_2 = 100
# X = np.r_[1.5 * rng.randn(n_samples_1, 2), 0.5 *
#           rng.randn(n_samples_2, 2) + [2, 2]]
# y = [0] * (n_samples_1) + [1] * (n_samples_2)
# print(X)
# print(y)

# # fit the model and get the separating hyperplane
# clf = svm.SVC(kernel='linear', C=1.0)
# clf.fit(X, y)

# w = clf.coef_[0]
# a = -w[0] / w[1]  # a可以理解为斜率
# xx = np.linspace(-5, 5)
# yy = a * xx - clf.intercept_[0] / w[1]  # 二维坐标下的直线方程


# # get the separating hyperplane using weighted classes
# wclf = svm.SVC(kernel='linear', class_weight={1: 10})
# wclf.fit(X, y)

# ww = wclf.coef_[0]
# wa = -ww[0] / ww[1]
# wyy = wa * xx - wclf.intercept_[0] / ww[1]  # 带权重的直线

# # plot separating hyperplanes and samples
# h0 = plt.plot(xx, yy, 'k-', label='no weights')
# h1 = plt.plot(xx, wyy, 'k--', label='with weights')
# plt.scatter(X[:, 0], X[:, 1], c=y)
# plt.legend()

# plt.axis('tight')
# plt.show()

# import numpy as np
# from sklearn.svm import SVR
# import matplotlib.pyplot as plt

# ###############################################################################
# # Generate sample data
# # 产生40组数据，每组一个数据，axis=0决定按列排列，=1表示行排列
# X = np.sort(5 * np.random.rand(40, 1), axis=0)
# y = np.sin(X).ravel()  # np.sin()输出的是列，和X对应，ravel表示转换成行

# ###############################################################################
# # Add noise to targets
# y[::5] += 3 * (0.5 - np.random.rand(8))

# ###############################################################################
# # Fit regression model
# svr_rbf = SVR(kernel='rbf', C=1e3, gamma=0.1)
# svr_lin = SVR(kernel='linear', C=1e3)
# svr_poly = SVR(kernel='poly', C=1e3, degree=2)
# y_rbf = svr_rbf.fit(X, y).predict(X)
# y_lin = svr_lin.fit(X, y).predict(X)
# y_poly = svr_poly.fit(X, y).predict(X)

# ###############################################################################
# # look at the results
# lw = 2
# plt.scatter(X, y, color='darkorange', label='data')
# plt.hold('on')
# plt.plot(X, y_rbf, color='navy', lw=lw, label='RBF model')
# plt.plot(X, y_lin, color='c', lw=lw, label='Linear model')
# plt.plot(X, y_poly, color='cornflowerblue', lw=lw, label='Polynomial model')
# plt.xlabel('data')
# plt.ylabel('target')
# plt.title('Support Vector Regression')
# plt.legend()
# plt.show()

import csv
import codecs

#较老excel版本，尾缀为xls
import xlrd
import xlwt

#处理xlsx文件
from openpyxl import load_workbook

filename = '/home/yaya/Downloads/百分比修改.xlsx'
wb = load_workbook(filename)
'''
#获取所有表格(worksheet)的名字
sheets = wb.get_sheet_names()
#第一个表格的名称
sheet_first = sheets[0]
#获取特定的worksheet
ws = wb.get_sheet_by_name(sheet_first)
'''
ws = wb.get_sheet_by_name('Sheet1')
#获取表格所有行和列，两者都是可迭代的
rows = ws.rows
columns = ws.columns
#迭代所有的行
datas = []
for row in rows:
    line = [col.value for col in row]
    datas.append(line)

#通过坐标读取值
# print(ws['B3'].value)
# print(ws.cell(row=1, column=1).value)
from sklearn import svm
import numpy as np
import matplotlib.pyplot as plt

# print(np.linspace(-1,3.5))

from sklearn.datasets.samples_generator import make_blobs
X, y = make_blobs(n_samples=50, centers=2,
                  random_state=0, cluster_std=0.60)

# print(X[0][0])
# print(X[:, 0])
print(type(y[0]))
# print(X[:, 1])
y = np.array(list(map(lambda x: x[0], datas)))
print(type(y[0]))
X = np.array(datas)
plt.scatter(X[:,7],X[:,2],c = y, s=5, cmap="spring")
plt.scatter(X[:, 7], X[:, 3], c=y, s=5, cmap="flag")
plt.scatter(X[:, 7], X[:, 4], c=y, s=5, cmap="RdBu")
plt.scatter(X[:, 7], X[:, 5], c=y, s=5, cmap="jet")
plt.scatter(X[:, 7], X[:, 6], c=y, s=5, cmap="winter")
plt.show()
